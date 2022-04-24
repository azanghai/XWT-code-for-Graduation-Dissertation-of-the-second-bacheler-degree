# code by azanghai
import json
import openpyxl
import requests
import os
import time
from cnsenti import Sentiment
from cnsenti import Emotion
import sys

sys.path.append('\\weibo-clawer')
from weibocrawler.weibo import main as calwer_user_weibo
import pandas as pd


# current access token: ''
# weibo_auth pre set required.
def get_comment(give_access_token, give_id, give_since_id=None, give_max_id=None, give_count=None, give_page=None):
    url = 'https://api.weibo.com/2/comments/show.json'

    # access_token : weibo_auth/UserSinaweibopy3.py 中运行，手动复制在access_token，以''形式传入
    # id : 需要查询评论的微博id，可在https://m.weibo.cn/中获得，以int形式传入
    # since_id 若指定此参数，则返回ID比since_id大的评论（即比since_id时间晚的评论），默认为0。具体时间形式未尝试,以int形式传入
    # max_id 若指定此参数，则返回ID小于或等于max_id的评论，默认为0。以int形式传入
    # count 单页返回的记录条数，默认为50。 int
    # page 返回结果的页码，默认为1。 int
    # filter_by_author 这个参数我没往里面写，作者筛选类型，0：全部、1：我关注的人、2：陌生人，默认为0。 int
    # 返回的内容是json格式，return解析为json格式

    paramas = {'access_token': give_access_token,
               'id': give_id,
               'since_id': give_since_id,
               'max_id': give_max_id,
               'count': give_count,
               'page': give_page}

    response = requests.get(url, params=paramas)
    if response.status_code == 200:
        return response.json()
    else:
        statuscondiction = 000
        print(response.text)
        return statuscondiction


def write_comment(data, filename, give_ID, page, ):
    if data == 000:
        print('data error!')
    else:
        path = '.\comment_jsondata\\'
        data_path1 = '.\\' + str(filename) + '\\'
        data_path2 = str(filename) + '\\'

        # set saving dir
        try:
            os.makedirs(path)
        except:
            pass
        try:
            os.makedirs(data_path1)
        except:
            pass

        # set workbook names
        workbookname = str(give_ID) + '_' + 'comment_data.xlsx'
        # worksheetname = str(give_ID)

        # headings of excel
        headings = ['form_weibo_id',
                    'user_id',
                    'user_name',
                    'comment_date',
                    'comment_content',
                    'user_display_location',
                    'user_description',
                    'user_diaplay_gender',
                    'user_folloewrs',
                    'user_watchcount',
                    'user_weibo_count',
                    'user_account_created_at',
                    'able_to_msg',
                    'is_bigV']

        # write raw data to txt
        direct_path_raw = path + '_' + str(page) + '_' + filename + str(give_ID) + '.txt'
        with open(direct_path_raw, encoding='utf-8', mode='w') as f:
            f.write(json.dumps(data))

        # form col lists to add cols easier in the future
        from_weibo_id = []
        user_id = []
        user_name = []
        comment_date = []
        comment = []
        location = []
        description = []
        gender = []
        follower = []
        watch = []
        weibonum = []
        created_at = []
        msg = []
        bigv = []

        # get information
        for i in data['comments']:
            from_weibo_id.append(str(give_ID))
            user_id.append(i['user']['id'])
            user_name.append(i['user']['screen_name'])
            comment_date.append(i['created_at'])
            comment.append(i['text'])
            location.append(i['user']['location'])
            description.append(i['user']['description'])
            gender.append(i['user']['gender'])
            follower.append(i['user']['followers_count'])
            watch.append(i['user']['friends_count'])
            weibonum.append(i['user']['statuses_count'])
            created_at.append(i['user']['created_at'])
            msg.append(i['user']['allow_all_act_msg'])
            bigv.append(i['user']['verified'])

        # creat workbook
        if os.path.exists(data_path2 + workbookname) == False:
            wb = openpyxl.Workbook()
            wb.save(filename=(data_path2 + workbookname))
        else:
            pass

        # read workbook to write
        wb = openpyxl.load_workbook(filename=(data_path2 + workbookname))
        try:
            ws = wb.active()
        except:
            pass
        sheet1 = wb.worksheets[0]

        # write headings
        if sheet1.cell(1, 1).value == 'form_weibo_id':
            pass
        else:
            for k in range(len(headings)):
                sheet1.cell(1, k + 1).value = headings[k]

        # read lines to decide where to write
        row_len = sheet1.max_row
        print('there is ' + str(row_len - 1) + ' comment in total.')

        for j in range(len(user_id)):
            sheet1.cell(row_len + j + 1, 1).value = from_weibo_id[j]
            sheet1.cell(row_len + j + 1, 2).value = user_id[j]
            sheet1.cell(row_len + j + 1, 3).value = user_name[j]
            sheet1.cell(row_len + j + 1, 4).value = comment_date[j]
            sheet1.cell(row_len + j + 1, 5).value = comment[j]
            sheet1.cell(row_len + j + 1, 6).value = location[j]
            sheet1.cell(row_len + j + 1, 7).value = description[j]
            sheet1.cell(row_len + j + 1, 8).value = gender[j]
            sheet1.cell(row_len + j + 1, 9).value = follower[j]
            sheet1.cell(row_len + j + 1, 10).value = watch[j]
            sheet1.cell(row_len + j + 1, 11).value = weibonum[j]
            sheet1.cell(row_len + j + 1, 12).value = created_at[j]
            sheet1.cell(row_len + j + 1, 13).value = msg[j]
            sheet1.cell(row_len + j + 1, 14).value = bigv[j]
        wb.save(filename=(data_path2 + workbookname))
        return_path = data_path2 + workbookname
    return return_path


def emotion_analysis(file_path, ):
    # load workbook and work sheet
    wb = openpyxl.load_workbook(file_path)
    st = wb.worksheets[0]

    # prepare list to read comment and result
    comment_list = []
    emotion_result = []
    senti_count_result = []
    senti_caluate_result = []

    # iteration to get comment(first line included)
    # it will find comment to collect
    letters = "abcdefghijklmnopqrstuvwxyz"
    for i in st['1']:
        # get comment target col number
        if i.value == 'comment_content':
            target_col = i.column
            # add value if matches
            for cell in st[letters[target_col - 1]]:
                comment_list.append(cell.value)

    if len(comment_list) > 1:

        # generate emotion result
        for i in comment_list:
            emotion = Emotion()
            emotion_result.append(emotion.emotion_count(i))
            senti = Sentiment()
            senti_count_result.append(senti.sentiment_count(i))
            senti_caluate_result.append(senti.sentiment_calculate(i))

        # write result
        # get pre write col
        pre_write_col = st.max_column

        # write calculate data(first line included)
        for i in range(len(emotion_result)):
            st.cell(i + 1, pre_write_col + 1).value = emotion_result[i]['words']
            st.cell(i + 1, pre_write_col + 2).value = emotion_result[i]['sentences']
            st.cell(i + 1, pre_write_col + 3).value = emotion_result[i]['好']
            st.cell(i + 1, pre_write_col + 4).value = emotion_result[i]['乐']
            st.cell(i + 1, pre_write_col + 5).value = emotion_result[i]['哀']
            st.cell(i + 1, pre_write_col + 6).value = emotion_result[i]['怒']
            st.cell(i + 1, pre_write_col + 7).value = emotion_result[i]['惧']
            st.cell(i + 1, pre_write_col + 8).value = emotion_result[i]['恶']
            st.cell(i + 1, pre_write_col + 9).value = emotion_result[i]['惊']
            st.cell(i + 1, pre_write_col + 10).value = senti_count_result[i]['pos']
            st.cell(i + 1, pre_write_col + 11).value = senti_count_result[i]['neg']
            st.cell(i + 1, pre_write_col + 12).value = senti_caluate_result[i]['pos']
            st.cell(i + 1, pre_write_col + 13).value = senti_caluate_result[i]['neg']

        # write first line
        st.cell(1, pre_write_col + 1).value = 'words'
        st.cell(1, pre_write_col + 2).value = 'sentences'
        st.cell(1, pre_write_col + 3).value = 'hao'
        st.cell(1, pre_write_col + 4).value = 'le'
        st.cell(1, pre_write_col + 5).value = 'ai'
        st.cell(1, pre_write_col + 6).value = 'nu'
        st.cell(1, pre_write_col + 7).value = 'ju'
        st.cell(1, pre_write_col + 8).value = 'wu'
        st.cell(1, pre_write_col + 9).value = 'jing'
        st.cell(1, pre_write_col + 10).value = 'count_pos'
        st.cell(1, pre_write_col + 11).value = 'count_neg'
        st.cell(1, pre_write_col + 12).value = 'calculate_pos'
        st.cell(1, pre_write_col + 13).value = 'calculate_neg'
        wb.save(filename=file_path)
    else:
        print('give up analysis, due to comment data error!')


def step_1_form_comment_excel(ID_list, give_acess_token):
    # current token: ''
    # input page number needed
    page_require = 10  # change number here,max=10
    # ID_list = []
    N = page_require + 1
    for id_num in ID_list:
        for pgnum in range(1, N):
            a = get_comment(give_access_token=give_acess_token,
                            give_id=id_num,
                            give_count=200,
                            give_page=pgnum
                            )
            print(a)
            print(type(a))
            b = write_comment(data=a, filename='21_5_8_ten_weibo', give_ID=id_num, page=pgnum)
            print('page ' + str(pgnum) + ' is complete!')
            time.sleep(20)
        emotion_analysis(b)
        print('complete!')

        time.sleep(35)


def step_2_claw_user_weibo():
    # file pre set required
    calwer_user_weibo()


def step_3_rewrite_file_to_excel_with_emotion_result(file_dir='', to_file_dir=''):
    def emotion_analysis(file_path):
        # load workbook and work sheet
        wb = openpyxl.load_workbook(file_path)
        st = wb.worksheets[0]

        # prepare list to read comment and result
        comment_list = []
        emotion_result = []
        senti_count_result = []
        senti_caluate_result = []

        # iteration to get comment(first line included)
        # it will find comment to collect
        letters = "abcdefghijklmnopqrstuvwxyz"
        for i in st['1']:
            # get comment target col number
            if i.value == '正文':
                target_col = i.column
                # add value if matches
                for cell in st[letters[target_col - 1]]:
                    comment_list.append(cell.value)

        if len(comment_list) > 1:

            # generate emotion result
            for i in comment_list:
                emotion = Emotion()
                emotion_result.append(emotion.emotion_count(i))
                senti = Sentiment()
                senti_count_result.append(senti.sentiment_count(i))
                senti_caluate_result.append(senti.sentiment_calculate(i))

            # write result
            # get pre write col
            pre_write_col = st.max_column

            # write calculate data(first line included)
            for i in range(len(emotion_result)):
                st.cell(i + 1, pre_write_col + 1).value = emotion_result[i]['words']
                st.cell(i + 1, pre_write_col + 2).value = emotion_result[i]['sentences']
                st.cell(i + 1, pre_write_col + 3).value = emotion_result[i]['好']
                st.cell(i + 1, pre_write_col + 4).value = emotion_result[i]['乐']
                st.cell(i + 1, pre_write_col + 5).value = emotion_result[i]['哀']
                st.cell(i + 1, pre_write_col + 6).value = emotion_result[i]['怒']
                st.cell(i + 1, pre_write_col + 7).value = emotion_result[i]['惧']
                st.cell(i + 1, pre_write_col + 8).value = emotion_result[i]['恶']
                st.cell(i + 1, pre_write_col + 9).value = emotion_result[i]['惊']
                st.cell(i + 1, pre_write_col + 10).value = senti_count_result[i]['pos']
                st.cell(i + 1, pre_write_col + 11).value = senti_count_result[i]['neg']
                st.cell(i + 1, pre_write_col + 12).value = senti_caluate_result[i]['pos']
                st.cell(i + 1, pre_write_col + 13).value = senti_caluate_result[i]['neg']

            # write first line
            st.cell(1, pre_write_col + 1).value = 'words'
            st.cell(1, pre_write_col + 2).value = 'sentences'
            st.cell(1, pre_write_col + 3).value = 'hao'
            st.cell(1, pre_write_col + 4).value = 'le'
            st.cell(1, pre_write_col + 5).value = 'ai'
            st.cell(1, pre_write_col + 6).value = 'nu'
            st.cell(1, pre_write_col + 7).value = 'ju'
            st.cell(1, pre_write_col + 8).value = 'wu'
            st.cell(1, pre_write_col + 9).value = 'jing'
            st.cell(1, pre_write_col + 10).value = 'count_pos'
            st.cell(1, pre_write_col + 11).value = 'count_neg'
            st.cell(1, pre_write_col + 12).value = 'calculate_pos'
            st.cell(1, pre_write_col + 13).value = 'calculate_neg'
            wb.save(filename=file_path)
        else:
            print('give up analysis, due to comment data error!')

    for root, dirs, files in os.walk(file_dir):
        for file in files:
            # 获取文件路径
            try:
                os.makedirs(to_file_dir)
            except:
                print('路径存在无需新建')
            print(file)

            data_file_path = os.path.join(root, file)
            print(data_file_path)
            try:
                csv = pd.read_csv(data_file_path, encoding='utf-8')
                excel_file_name = to_file_dir + '\\' + file.strip('.csv') + '.xlsx'
                print(excel_file_name)
                csv.to_excel(excel_file_name, sheet_name='data')
                emotion_analysis(excel_file_name)
            except:
                print('error')
                print(data_file_path)


def step_4_write_summary_excel(user_detail_file='.\emotion_weibo_216', user_chart_detail_path='216pilot.xlsx',
                               destfile='FINAL_CHART_216.xlsx'):
    input('press any key')
    summary_workbook = openpyxl.load_workbook(user_detail_file + '\\user.xlsx')
    st1 = summary_workbook.worksheets[0]
    All_data_list = []
    for i in range(0, st1.max_row):
        All_data_list.append({'i': ''})
    temp_dict = {'userid': '',
                 'user_name': '',
                 'user_all_weibo_num': '',
                 'user_follower': '',
                 'user_watch': '',
                 'user_weibo_rank': '',
                 'gender': '',
                 'age': '',
                 'extraversion': '',
                 'agreeableness': '',
                 'conscientiousness': '',
                 'neuroticism': '',
                 'openness': '',
                 'bsmas': '',
                 'uls-6': '',
                 'stress': '',
                 'anxiety': '',
                 'depression': '',
                 'claw_weibo_num': '',
                 'pos_sum': '',
                 'neg_sum': '',
                 'hao': '',
                 'le': '',
                 'ai': '',
                 'nu': '',
                 'ju': '',
                 'wu': '',
                 'jing': '',
                 'take_time': '',
                 'neg_big_pos': '',
                 'pos_big_neg': '',
                 'neg_euqal_pos': ''
                 }
    for i in range(2, st1.max_row + 1):
        require_col = [2, 3, 11, 12, 13, 18]
        All_data_list[i - 2]['userid'] = st1.cell(i, 2).value
        All_data_list[i - 2]['user_name'] = st1.cell(i, 3).value
        All_data_list[i - 2]['user_all_weibo_num'] = st1.cell(i, 11).value
        All_data_list[i - 2]['user_follower'] = st1.cell(i, 12).value
        All_data_list[i - 2]['user_watch'] = st1.cell(i, 13).value
        All_data_list[i - 2]['user_weibo_rank'] = st1.cell(i, 18).value

    summary_workbook.close()

    for single_dict in All_data_list:
        try:

            user_all_data_file = user_detail_file + '\\' + str(single_dict['userid']) + '.xlsx'
            user_all_data_workbook = openpyxl.load_workbook(user_all_data_file)
            st2 = user_all_data_workbook.worksheets[0]
            require_col = [27, 28, 18, 19, 20, 21, 22, 23, 24]
            pos_sum = 0
            neg_sum = 0
            hao_sum = 0
            le_sum = 0
            ai_sum = 0
            nu_sum = 0
            ju_sum = 0
            wu_sum = 0
            jing_sum = 0
            for i in range(2, st2.max_row + 1):
                pos_sum = pos_sum + int(st2.cell(i, 27).value)
                neg_sum = neg_sum + int(st2.cell(i, 28).value)
                hao_sum = hao_sum + int(st2.cell(i, 18).value)
                le_sum = le_sum + int(st2.cell(i, 19).value)
                ai_sum = ai_sum + int(st2.cell(i, 20).value)
                nu_sum = nu_sum + int(st2.cell(i, 21).value)
                ju_sum = ju_sum + int(st2.cell(i, 22).value)
                wu_sum = wu_sum + int(st2.cell(i, 23).value)
                jing_sum = jing_sum + int(st2.cell(i, 24).value)

            single_dict['pos_sum'] = pos_sum
            single_dict['neg_sum'] = neg_sum
            single_dict['hao'] = hao_sum
            single_dict['le'] = le_sum
            single_dict['ai'] = ai_sum
            single_dict['nu'] = nu_sum
            single_dict['ju'] = ju_sum
            single_dict['wu'] = wu_sum
            single_dict['jing'] = jing_sum
            single_dict['claw_weibo_num'] = st2.max_row - 1

            neg_big_pos = 0
            pos_big_neg = 0
            neg_euqal_pos = 0
            for i in range(2, st2.max_row + 1):
                # print(float(st2.cell(i, 28).value))
                # print(float(st2.cell(i, 27).value))
                if float(st2.cell(i, 28).value) > float(st2.cell(i, 27).value):
                    neg_big_pos = neg_big_pos + 1
                elif float(st2.cell(i, 28).value) == float(st2.cell(i, 27).value):
                    neg_euqal_pos = neg_euqal_pos + 1
                elif float(st2.cell(i, 28).value) < float(st2.cell(i, 27).value):
                    pos_big_neg = pos_big_neg + 1
                else:
                    print('i have no idea about the relation between neg and pos in {} '.format(single_dict['userid']))
            single_dict['neg_big_pos'] = neg_big_pos
            single_dict['pos_big_neg'] = pos_big_neg
            single_dict['neg_euqal_pos'] = neg_euqal_pos

        except:
            try:
                print(str(single_dict['userid']) + '这个没爬到')
            except:
                pass

    summary_workbook.close()

    user_chart_detail_workbook = openpyxl.load_workbook(user_chart_detail_path)
    st3 = user_chart_detail_workbook.worksheets[0]
    for single_dict in All_data_list:
        try:
            for i in range(2, st3.max_row + 1):
                if int(st3.cell(i, 3).value) == int(single_dict['userid']):
                    single_dict['gender'] = st3.cell(i, 4).value
                    single_dict['age'] = st3.cell(i, 5).value
                    single_dict['extraversion'] = st3.cell(i, 7).value
                    single_dict['agreeableness'] = st3.cell(i, 8).value
                    single_dict['conscientiousness'] = st3.cell(i, 9).value
                    single_dict['neuroticism'] = st3.cell(i, 10).value
                    single_dict['openness'] = st3.cell(i, 11).value
                    single_dict['bsmas'] = st3.cell(i, 12).value
                    single_dict['uls-6'] = st3.cell(i, 13).value
                    single_dict['stress'] = st3.cell(i, 14).value
                    single_dict['anxiety'] = st3.cell(i, 15).value
                    single_dict['depression'] = st3.cell(i, 16).value
                    single_dict['take_time'] = st3.cell(i, 2).value
        except:
            pass
    user_chart_detail_workbook.close()

    final_workbook = openpyxl.Workbook()
    st4 = final_workbook.worksheets[0]
    headings = ['用户id',
                '用户昵称',
                '全微博数',
                '粉丝',
                '关注',
                '微博等级',
                '性别',
                '年龄',
                '外倾性',
                '宜人性',
                '责任心',
                '神经质性',
                '开放性',
                '社交媒体成瘾',
                '孤独感',
                '压力',
                '焦虑',
                '抑郁',
                '爬取微博数',
                '总计积极分',
                '总计消极分',
                '总计好',
                '总计乐',
                '总计哀',
                '总计怒',
                '总计惧',
                '总计恶',
                '总计惊',
                '总用时',
                '消极更大数量',
                '积极更大数量',
                '相等数量', ]
    for i in range(len(headings)):
        st4.cell(1, i + 1).value = headings[i]
    for i in range(2, len(All_data_list) + 1):
        try:
            st4.cell(i, 1).value = All_data_list[i - 2]['userid']
            st4.cell(i, 2).value = All_data_list[i - 2]['user_name']
            st4.cell(i, 3).value = All_data_list[i - 2]['user_all_weibo_num']
            st4.cell(i, 4).value = All_data_list[i - 2]['user_follower']
            st4.cell(i, 5).value = All_data_list[i - 2]['user_watch']
            st4.cell(i, 6).value = All_data_list[i - 2]['user_weibo_rank']
            st4.cell(i, 7).value = All_data_list[i - 2]['gender']
            st4.cell(i, 8).value = All_data_list[i - 2]['age']
            st4.cell(i, 9).value = All_data_list[i - 2]['extraversion']
            st4.cell(i, 10).value = All_data_list[i - 2]['agreeableness']
            st4.cell(i, 11).value = All_data_list[i - 2]['conscientiousness']
            st4.cell(i, 12).value = All_data_list[i - 2]['neuroticism']
            st4.cell(i, 13).value = All_data_list[i - 2]['openness']
            st4.cell(i, 14).value = All_data_list[i - 2]['bsmas']
            st4.cell(i, 15).value = All_data_list[i - 2]['uls-6']
            st4.cell(i, 16).value = All_data_list[i - 2]['stress']
            st4.cell(i, 17).value = All_data_list[i - 2]['anxiety']
            st4.cell(i, 18).value = All_data_list[i - 2]['depression']
            st4.cell(i, 19).value = All_data_list[i - 2]['claw_weibo_num']
            st4.cell(i, 20).value = All_data_list[i - 2]['pos_sum']
            st4.cell(i, 21).value = All_data_list[i - 2]['neg_sum']
            st4.cell(i, 22).value = All_data_list[i - 2]['hao']
            st4.cell(i, 23).value = All_data_list[i - 2]['le']
            st4.cell(i, 24).value = All_data_list[i - 2]['ai']
            st4.cell(i, 25).value = All_data_list[i - 2]['nu']
            st4.cell(i, 26).value = All_data_list[i - 2]['ju']
            st4.cell(i, 27).value = All_data_list[i - 2]['wu']
            st4.cell(i, 28).value = All_data_list[i - 2]['jing']
            st4.cell(i, 29).value = All_data_list[i - 2]['take_time']
            st4.cell(i, 30).value = All_data_list[i - 2]['neg_big_pos']
            st4.cell(i, 31).value = All_data_list[i - 2]['pos_big_neg']
            st4.cell(i, 32).value = All_data_list[i - 2]['neg_equal_pos']
        except:
            pass
    final_workbook.save(filename=destfile)


if __name__ == '__main__':
    step_3_rewrite_file_to_excel_with_emotion_result()
    step_4_write_summary_excel()
