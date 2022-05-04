# code by zanghai

import json
import openpyxl
import requests
import os
import time
from cnsenti import Sentiment
from cnsenti import Emotion


# current access token: ''
def get_comment(give_access_token, give_id, give_since_id=None, give_max_id=None, give_count=None, give_page=None):
    url = 'https://api.weibo.com/2/comments/show.json'

    # access_token : weibo_auth/UserSinaweibopy3.py 中运行，手动复制在access_token，以''形式传入
    # id : 需要查询评论的微博id，可在https://m.weibo.cn/中获得，以int形式传入
    # since_id 若指定此参数，则返回ID比since_id大的评论（即比since_id时间晚的评论），默认为0。具体时间形式未尝试,以int形式传入
    # max_id 若指定此参数，则返回ID小于或等于max_id的评论，默认为0。以int形式传入
    # count 单页返回的记录条数，默认为50。 int
    # page 返回结果的页码，默认为1。 int
    # filter_by_author 这个参数我没往里面写，作者筛选类型，0：全部、1：我关注的人、2：陌生人，默认为0。 int
    # 返回的内容是json格式，return解析为json格式

    paramas = {'access_token': give_access_token,
               'id': give_id,
               'since_id': give_since_id,
               'max_id': give_max_id,
               'count': give_count,
               'page': give_page}

    response = requests.get(url, params=paramas)
    if response.status_code == 200:
        return response.json()
    else:
        statuscondiction = 000
        print(response.text)
        return statuscondiction


def write_comment(data, filename, give_ID, page, ):
    if data == 000:
        print('data error!')
    else:
        path = '.\jsondata\\'
        data_path1 = '.\\' + str(filename) + '\\'
        data_path2 = str(filename) + '\\'

        # set saving dir
        try:
            os.makedirs(path)
        except:
            pass
        try:
            os.makedirs(data_path1)
        except:
            pass

        # set workbook names
        workbookname = str(give_ID) + '_' + 'w_c_r.xlsx'
        # worksheetname = str(give_ID)

        # headings of excel
        headings = ['form_weibo_id',
                    'user_id',
                    'user_name',
                    'comment_date',
                    'comment_content',
                    'user_display_location',
                    'user_description',
                    'user_diaplay_gender',
                    'user_folloewrs',
                    'user_watchcount',
                    'user_weibo_count',
                    'user_account_created_at',
                    'able_to_msg',
                    'is_bigV']

        # write raw data to txt
        direct_path_raw = path + '_' + str(page) + '_' + filename + str(give_ID) + '.txt'
        with open(direct_path_raw, encoding='utf-8', mode='w') as f:
            f.write(json.dumps(data))

        # form col lists to add cols easier in the future
        from_weibo_id = []
        user_id = []
        user_name = []
        comment_date = []
        comment = []
        location = []
        description = []
        gender = []
        follower = []
        watch = []
        weibonum = []
        created_at = []
        msg = []
        bigv = []

        # get information
        for i in data['comments']:
            from_weibo_id.append(str(give_ID))
            user_id.append(i['user']['id'])
            user_name.append(i['user']['screen_name'])
            comment_date.append(i['created_at'])
            comment.append(i['text'])
            location.append(i['user']['location'])
            description.append(i['user']['description'])
            gender.append(i['user']['gender'])
            follower.append(i['user']['followers_count'])
            watch.append(i['user']['friends_count'])
            weibonum.append(i['user']['statuses_count'])
            created_at.append(i['user']['created_at'])
            msg.append(i['user']['allow_all_act_msg'])
            bigv.append(i['user']['verified'])

        # creat workbook
        if os.path.exists(data_path2 + workbookname) == False:
            wb = openpyxl.Workbook()
            wb.save(filename=(data_path2 + workbookname))
        else:
            pass

        # read workbook to write
        wb = openpyxl.load_workbook(filename=(data_path2 + workbookname))
        try:
            ws = wb.active()
        except:
            pass
        sheet1 = wb.worksheets[0]

        # write headings
        if sheet1.cell(1, 1).value == 'form_weibo_id':
            pass
        else:
            for k in range(len(headings)):
                sheet1.cell(1, k + 1).value = headings[k]

        # read lines to decide where to write
        row_len = sheet1.max_row
        print('there is ' + str(row_len - 1) + ' comment in total.')

        for j in range(len(user_id)):
            sheet1.cell(row_len + j + 1, 1).value = from_weibo_id[j]
            sheet1.cell(row_len + j + 1, 2).value = user_id[j]
            sheet1.cell(row_len + j + 1, 3).value = user_name[j]
            sheet1.cell(row_len + j + 1, 4).value = comment_date[j]
            sheet1.cell(row_len + j + 1, 5).value = comment[j]
            sheet1.cell(row_len + j + 1, 6).value = location[j]
            sheet1.cell(row_len + j + 1, 7).value = description[j]
            sheet1.cell(row_len + j + 1, 8).value = gender[j]
            sheet1.cell(row_len + j + 1, 9).value = follower[j]
            sheet1.cell(row_len + j + 1, 10).value = watch[j]
            sheet1.cell(row_len + j + 1, 11).value = weibonum[j]
            sheet1.cell(row_len + j + 1, 12).value = created_at[j]
            sheet1.cell(row_len + j + 1, 13).value = msg[j]
            sheet1.cell(row_len + j + 1, 14).value = bigv[j]
        wb.save(filename=(data_path2 + workbookname))
        return_path = data_path2 + workbookname
    return return_path


def emotion_analysis(file_path, ):
    # load workbook and work sheet
    wb = openpyxl.load_workbook(file_path)
    st = wb.worksheets[0]

    # prepare list to read comment and result
    comment_list = []
    emotion_result = []
    senti_count_result = []
    senti_caluate_result = []

    # iteration to get comment(first line included)
    # it will find comment to collect
    letters = "abcdefghijklmnopqrstuvwxyz"
    for i in st['1']:
        # get comment target col number
        if i.value == 'comment_content':
            target_col = i.column
            # add value if matches
            for cell in st[letters[target_col - 1]]:
                comment_list.append(cell.value)

    if len(comment_list) > 1:

        # generate emotion result
        for i in comment_list:
            emotion = Emotion()
            emotion_result.append(emotion.emotion_count(i))
            senti = Sentiment()
            senti_count_result.append(senti.sentiment_count(i))
            senti_caluate_result.append(senti.sentiment_calculate(i))

        # write result
        # get pre write col
        pre_write_col = st.max_column

        # write calculate data(first line included)
        for i in range(len(emotion_result)):
            st.cell(i + 1, pre_write_col + 1).value = emotion_result[i]['words']
            st.cell(i + 1, pre_write_col + 2).value = emotion_result[i]['sentences']
            st.cell(i + 1, pre_write_col + 3).value = emotion_result[i]['好']
            st.cell(i + 1, pre_write_col + 4).value = emotion_result[i]['乐']
            st.cell(i + 1, pre_write_col + 5).value = emotion_result[i]['哀']
            st.cell(i + 1, pre_write_col + 6).value = emotion_result[i]['怒']
            st.cell(i + 1, pre_write_col + 7).value = emotion_result[i]['惧']
            st.cell(i + 1, pre_write_col + 8).value = emotion_result[i]['恶']
            st.cell(i + 1, pre_write_col + 9).value = emotion_result[i]['惊']
            st.cell(i + 1, pre_write_col + 10).value = senti_count_result[i]['pos']
            st.cell(i + 1, pre_write_col + 11).value = senti_count_result[i]['neg']
            st.cell(i + 1, pre_write_col + 12).value = senti_caluate_result[i]['pos']
            st.cell(i + 1, pre_write_col + 13).value = senti_caluate_result[i]['neg']

        # write first line
        st.cell(1, pre_write_col + 1).value = 'words'
        st.cell(1, pre_write_col + 2).value = 'sentences'
        st.cell(1, pre_write_col + 3).value = 'hao'
        st.cell(1, pre_write_col + 4).value = 'le'
        st.cell(1, pre_write_col + 5).value = 'ai'
        st.cell(1, pre_write_col + 6).value = 'nu'
        st.cell(1, pre_write_col + 7).value = 'ju'
        st.cell(1, pre_write_col + 8).value = 'wu'
        st.cell(1, pre_write_col + 9).value = 'jing'
        st.cell(1, pre_write_col + 10).value = 'count_pos'
        st.cell(1, pre_write_col + 11).value = 'count_neg'
        st.cell(1, pre_write_col + 12).value = 'calculate_pos'
        st.cell(1, pre_write_col + 13).value = 'calculate_neg'
        wb.save(filename=file_path)
    else:
        print('give up analysis, due to comment data error!')

# 按间距中的绿色按钮以运行脚本。
# 当前可以获取某一微博下的评论，并将原始格式写入jsondata文件中，将处理后数据写入filename+微博id+wcr（weibo_comment_raw）.xlsx 文件中
# 并将情绪分析(七种基本情绪)\情感分析在后排输出(词库法)、可构建自己的词库（需要样本进行训练）
if __name__ == '__main__':
    # input page number needed
    page_require = 10  # change number here,max=10
    # ID = []
    ID = []

    N = page_require + 1
    for id_num in ID:
        for pgnum in range(1, N):
            a = get_comment(give_access_token='',
                            give_id=id_num,
                            give_count=200,
                            give_page=pgnum
                            )
            print(a)
            print(type(a))
            b = write_comment(data=a, filename='ligeweibo', give_ID=id_num, page=pgnum)
            print('page ' + str(pgnum) + ' is complete!')
            time.sleep(20)
        emotion_analysis(b)
        print('complete!')

        time.sleep(35)
